# -*- coding: utf-8 -*-
"""
Excel 工作表拆分导出工具
------------------------
功能：
  1. 拖拽或选择一个 Excel 文件（.xlsx / .xlsm）
  2. 显示工作簿中所有工作表，可勾选需要导出的工作表
  3. 选择导出目录
  4. 可选择是否将公式单元格转换为数值（保留格式，去除公式）
  5. 将每个勾选的工作表导出为单独的 Excel 文件（保留原有格式、合并单元格、列宽等）

依赖：
  pip install openpyxl tkinterdnd2

运行（Windows / macOS 通用）：
  python excel_sheet_splitter.py   （macOS 上用 python3）
"""

import os
import re
import sys
import threading
import urllib.parse
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import openpyxl
from openpyxl.cell.cell import MergedCell
try:
    from openpyxl.worksheet.formula import ArrayFormula
except ImportError:
    ArrayFormula = None

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    DND_AVAILABLE = True
except ImportError:
    DND_AVAILABLE = False
    TkinterDnD = None


class ExcelSplitterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel 工作表拆分导出工具")
        self.root.geometry("600x700")
        self.root.minsize(520, 520)

        self.excel_path = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.convert_formulas_var = tk.BooleanVar(value=False)
        self.sheet_vars = {}  # sheet_name -> BooleanVar
        self.workbook_path = None

        self._build_ui()

    # ---------------------------------------------------------------- UI ---
    def _build_ui(self):
        # 注意：下面先创建各区域内容，最后统一按顺序 pack 到窗口上。
        # 这样可以保证"导出目录"和"导出按钮"始终固定在窗口底部可见，
        # 不会被工作表列表（数量多时）挤到窗口外面看不到。

        # 1. 文件选择区
        file_frame = ttk.LabelFrame(self.root, text="1. 选择 Excel 文件", padding=10)

        self.drop_label = tk.Label(
            file_frame,
            text="将 Excel 文件拖到此处\n或点击下方按钮选择文件",
            relief="ridge",
            bd=2,
            height=4,
            bg="#f5f5f5",
            fg="#444",
        )
        self.drop_label.pack(fill="x", pady=(0, 8))

        btn_row = ttk.Frame(file_frame)
        btn_row.pack(fill="x")
        ttk.Button(btn_row, text="选择文件...", command=self.browse_file).pack(side="left")
        ttk.Label(btn_row, textvariable=self.excel_path, foreground="#555").pack(
            side="left", padx=10, fill="x", expand=True
        )

        if DND_AVAILABLE:
            self.drop_label.drop_target_register(DND_FILES)
            self.drop_label.dnd_bind("<<Drop>>", self.on_drop)
        else:
            self.drop_label.config(
                text="（拖拽功能不可用，请使用下方按钮选择文件）"
            )

        # 3. 导出目录区（先创建，稍后 pack 在窗口底部）
        out_frame = ttk.LabelFrame(self.root, text="3. 选择导出目录", padding=10)

        ttk.Button(out_frame, text="选择目录...", command=self.browse_output_dir).pack(side="left")
        ttk.Label(out_frame, textvariable=self.output_dir, foreground="#555").pack(
            side="left", padx=10, fill="x", expand=True
        )

        # 4. 导出选项区（先创建，稍后 pack 在窗口底部）
        options_frame = ttk.LabelFrame(self.root, text="4. 导出选项", padding=10)

        ttk.Checkbutton(
            options_frame,
            text="将公式单元格转换为数值（保留原有格式，只去除公式，固定结果）",
            variable=self.convert_formulas_var,
        ).pack(anchor="w")

        # 5. 导出按钮（先创建，稍后 pack 在窗口底部）
        action_frame = ttk.Frame(self.root)

        self.export_btn = ttk.Button(action_frame, text="导出选中的工作表", command=self.start_export)
        self.export_btn.pack(fill="x")

        # 6. 状态文字（先创建，稍后 pack 在窗口底部）
        self.status_label = ttk.Label(self.root, text="", foreground="#333")

        # 7. 作者签名（先创建，稍后 pack 在窗口最底部）
        import webbrowser

        row = ttk.Frame(self.root) 
        
        ttk.Label(
            row,
            text="插件制作者：筱理_Rize    ",
            foreground="#999",
            font=("TkDefaultFont", 9),
        ).pack(side="left")
        
        link_label = tk.Label(
            row,
            text="Project.ExcelSheetSplitter",                          # 换成你想要的文字
            fg="#4a90d9",
            font=("TkDefaultFont", 9, "underline"),
            cursor="hand2",
        )
        link_label.pack(side="left")
        link_label.bind("<Button-1>", lambda e: webbrowser.open("https://github.com/Rize-Nino/ExcelSheetSplitter"))

        row.pack(side="bottom", fill="x", padx=10, pady=(0, 6))

        # 2. 工作表勾选区（最后创建并 pack，占据中间剩余空间，内部可滚动）
        sheet_frame = ttk.LabelFrame(self.root, text="2. 勾选需要导出的工作表", padding=10)

        top_btns = ttk.Frame(sheet_frame)
        top_btns.pack(fill="x", pady=(0, 5))
        ttk.Button(top_btns, text="全选", command=self.select_all).pack(side="left")
        ttk.Button(top_btns, text="全不选", command=self.deselect_all).pack(side="left", padx=5)

        canvas_frame = ttk.Frame(sheet_frame)
        canvas_frame.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(canvas_frame, borderwidth=0, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=self.canvas.yview)
        self.checklist_frame = ttk.Frame(self.canvas)

        self.checklist_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")),
        )
        self.canvas.create_window((0, 0), window=self.checklist_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        # 鼠标滚轮支持（Windows/macOS 用 MouseWheel，Linux 用 Button-4/5）
        if sys.platform.startswith("linux"):
            self.canvas.bind_all("<Button-4>", self._on_mousewheel_linux)
            self.canvas.bind_all("<Button-5>", self._on_mousewheel_linux)
        else:
            self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        # ---- 统一按顺序 pack：顶部 -> 底部依次固定，中间区域自动填充 ----
        file_frame.pack(side="top", fill="x", padx=10, pady=(10, 5))
        self.status_label.pack(side="bottom", fill="x", padx=10, pady=(0, 4))
        action_frame.pack(side="bottom", fill="x", padx=10, pady=(4, 4))
        options_frame.pack(side="bottom", fill="x", padx=10, pady=(5, 0))
        out_frame.pack(side="bottom", fill="x", padx=10, pady=5)
        sheet_frame.pack(side="top", fill="both", expand=True, padx=10, pady=5)

    def _on_mousewheel(self, event):
        if sys.platform == "darwin":
            # macOS 触摸板/鼠标滚轮的 delta 单位与 Windows 不同，不需要除以120
            self.canvas.yview_scroll(int(-1 * event.delta), "units")
        else:
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _on_mousewheel_linux(self, event):
        delta = -1 if event.num == 4 else 1
        self.canvas.yview_scroll(delta, "units")

    # --------------------------------------------------------- 文件处理 ---
    def on_drop(self, event):
        # event.data 可能包含多个路径，路径含空格时会用 {} 包裹
        matches = re.findall(r"\{.*?\}|\S+", event.data)
        if not matches:
            return
        path = matches[0].strip("{}")
        if path.startswith("file://"):
            path = urllib.parse.unquote(urllib.parse.urlparse(path).path)
        self.load_file(path)

    def browse_file(self):
        path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")],
        )
        if path:
            self.load_file(path)

    def load_file(self, path):
        if not os.path.isfile(path):
            messagebox.showerror("错误", f"文件不存在：\n{path}")
            return
        if not path.lower().endswith((".xlsx", ".xlsm")):
            messagebox.showerror("错误", "请选择 .xlsx 或 .xlsm 格式的 Excel 文件")
            return

        self.status_label.config(text="正在读取工作簿...")
        self.root.update_idletasks()

        try:
            # 只读模式，仅用于快速获取工作表名称列表
            wb = openpyxl.load_workbook(path, read_only=True)
            sheet_names = list(wb.sheetnames)
            wb.close()
        except Exception as e:
            messagebox.showerror("错误", f"无法读取该 Excel 文件：\n{e}")
            self.status_label.config(text="")
            return

        self.excel_path.set(path)
        self.workbook_path = path
        self.populate_sheets(sheet_names)
        self.status_label.config(text=f"已加载，共 {len(sheet_names)} 个工作表")

        if not self.output_dir.get():
            self.output_dir.set(os.path.dirname(path))

    def populate_sheets(self, sheet_names):
        for child in self.checklist_frame.winfo_children():
            child.destroy()
        self.sheet_vars = {}

        for name in sheet_names:
            var = tk.BooleanVar(value=True)
            cb = ttk.Checkbutton(self.checklist_frame, text=name, variable=var)
            cb.pack(fill="x", anchor="w", pady=2)
            self.sheet_vars[name] = var

    def select_all(self):
        for var in self.sheet_vars.values():
            var.set(True)

    def deselect_all(self):
        for var in self.sheet_vars.values():
            var.set(False)

    # --------------------------------------------------------- 输出目录 ---
    def browse_output_dir(self):
        path = filedialog.askdirectory(title="选择导出目录")
        if path:
            self.output_dir.set(path)

    # --------------------------------------------------------------- 导出 ---
    def start_export(self):
        if not self.workbook_path:
            messagebox.showwarning("提示", "请先选择一个 Excel 文件")
            return

        selected = [name for name, var in self.sheet_vars.items() if var.get()]
        if not selected:
            messagebox.showwarning("提示", "请至少勾选一个工作表")
            return

        out_dir = self.output_dir.get()
        if not out_dir:
            messagebox.showwarning("提示", "请选择导出目录")
            return
        if not os.path.isdir(out_dir):
            try:
                os.makedirs(out_dir, exist_ok=True)
            except Exception as e:
                messagebox.showerror("错误", f"无法创建目录：\n{e}")
                return

        self.export_btn.config(state="disabled")
        self.status_label.config(text="正在导出，请稍候...")
        convert_formulas = self.convert_formulas_var.get()
        thread = threading.Thread(
            target=self._export_worker, args=(selected, out_dir, convert_formulas), daemon=True
        )
        thread.start()

    def _export_worker(self, selected, out_dir, convert_formulas):
        src_path = self.workbook_path
        is_xlsm = src_path.lower().endswith(".xlsm")
        errors = []
        done = 0

        wb_values = None
        if convert_formulas:
            try:
                # data_only=True 读取的是 Excel 上次保存时缓存的"计算结果"，
                # openpyxl 本身不会重新计算公式。如果源文件的公式从未被 Excel
                # 打开并保存过，缓存值可能为空。
                wb_values = openpyxl.load_workbook(src_path, data_only=True)
            except Exception as e:
                errors.append(f"读取公式缓存值失败：{e}")

        for sheet_name in selected:
            try:
                # 每个工作表都从原始文件重新加载一份完整工作簿，
                # 删除其余工作表后保存，从而完整保留该工作表的
                # 格式、合并单元格、列宽、条件格式等。
                wb = openpyxl.load_workbook(src_path, keep_vba=is_xlsm)

                if convert_formulas and wb_values is not None:
                    self._convert_formulas_to_values(wb[sheet_name], wb_values[sheet_name])

                for name in list(wb.sheetnames):
                    if name != sheet_name:
                        del wb[name]
                safe_name = self._sanitize_filename(sheet_name)
                ext = ".xlsm" if is_xlsm else ".xlsx"
                out_path = os.path.join(out_dir, f"{safe_name}{ext}")
                wb.save(out_path)
                wb.close()
                done += 1
            except Exception as e:
                errors.append(f"{sheet_name}: {e}")

        if wb_values is not None:
            wb_values.close()

        self.root.after(0, self._export_finished, done, errors, out_dir)

    @staticmethod
    def _convert_formulas_to_values(ws_formula, ws_value):
        """将 ws_formula 工作表中所有公式单元格的值替换为 ws_value 中对应的缓存计算结果。"""
        for row_f, row_v in zip(ws_formula.iter_rows(), ws_value.iter_rows()):
            for cell_f, cell_v in zip(row_f, row_v):
                if isinstance(cell_f, MergedCell):
                    continue
                value = cell_f.value
                is_formula = isinstance(value, str) and value.startswith("=")
                if not is_formula and ArrayFormula is not None:
                    is_formula = isinstance(value, ArrayFormula)
                if is_formula:
                    cell_f.value = cell_v.value

    def _export_finished(self, done, errors, out_dir):
        self.export_btn.config(state="normal")
        if errors:
            self.status_label.config(text=f"导出完成：成功 {done} 个，失败 {len(errors)} 个")
            messagebox.showerror("部分导出失败", "\n".join(errors))
        else:
            self.status_label.config(text=f"导出完成，共导出 {done} 个文件")
            messagebox.showinfo("完成", f"成功导出 {done} 个工作表文件到：\n{out_dir}")

    @staticmethod
    def _sanitize_filename(name):
        invalid_chars = '<>:"/\\|?*'
        for ch in invalid_chars:
            name = name.replace(ch, "_")
        name = name.strip()
        return name or "Sheet"


def main():
    global DND_AVAILABLE

    root = None
    if DND_AVAILABLE:
        try:
            root = TkinterDnD.Tk()
        except Exception as e:
            # 某些系统（常见于 macOS）上 tkinterdnd2 自带的原生 tkdnd 库
            # 与系统 Tcl/Tk 版本不兼容，会在这里抛出异常。出现这种情况时
            # 自动降级为不支持拖拽的普通窗口，而不是让程序直接崩溃退出，
            # 用户仍可以用"选择文件..."按钮正常使用其余全部功能。
            detail = str(e)
            underlying = e.__cause__ or e.__context__
            if underlying is not None and str(underlying) != detail:
                detail += f"｜底层原因：{type(underlying).__name__}: {underlying}"
            print(f"[提示] 拖拽支持组件初始化失败，已自动禁用拖拽功能（不影响其他功能）：{detail}")
            DND_AVAILABLE = False

            # TkinterDnD.Tk() 内部会先创建一个真正的 Tk 根窗口，再去加载
            # 拖拽库；如果加载失败，前面已创建的窗口不会被自动销毁，会变成
            # 一个孤立的空白窗口（标题通常显示为默认的 "tk"）留在屏幕上。
            # 这里把这个遗留窗口找出来并隐藏掉（用 withdraw 而不是 destroy，
            # 避免触发 Tcl 内部排队中的 <<ThemeChanged>> 等事件报错）。
            leftover = getattr(tk, "_default_root", None)
            if leftover is not None:
                try:
                    leftover.withdraw()
                except Exception:
                    pass
                tk._default_root = None

    if root is None:
        root = tk.Tk()

    try:
        style = ttk.Style(root)
        available = style.theme_names()
        if sys.platform == "win32" and "vista" in available:
            style.theme_use("vista")
        elif sys.platform == "darwin" and "aqua" in available:
            style.theme_use("aqua")
    except Exception:
        pass

    app = ExcelSplitterApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
